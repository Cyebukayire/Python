import openpyxl as xl
from openpyxl.chart import BarChart, Reference

workbook = xl.load_workbook("transactions.xlsx")
sheet1 = workbook['Sheet1']

# cell1 = sheet1['a1']
# cell1 = sheet1.cell(1, 23)
# print(cell1.value)

# print(sheet1.max_column)
# print(sheet1.max-row)

#  Print all the values
for row in range(2, sheet1.max_row + 1):
    cell = sheet1.cell(row, 3)
    correct_price = cell.value * 0.9
    correct_price_cell = sheet1.cell(row, 4)
    correct_price_cell.value = correct_price
sheet1.cell(1, 4).value = "Correct Price"

values = Reference(sheet1,
          min_row=2,
          max_row=sheet1.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet1.add_chart(chart, 'e2')

workbook.save("transactions2.xlsx")

